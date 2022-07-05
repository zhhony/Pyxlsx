import modules.pyappui as pyappui
import openpyxl
import xlrd
import xlwt
from PySide6.QtCore import *
from PySide6.QtGui import *
from PySide6.QtWidgets import *
from pathlib import Path
from typing import *
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.styles import Font
from py_test_tools import *


class DirPage():
    def __init__(self, MainWindow: QMainWindow, ui: pyappui.Ui_MainWindow) -> None:
        self._EXCEL_FILE_DIR = './'  # 默认的工作目录
        self._EXCEL_FLAG = None  # excel文件类型的标识，1标识xlsx，0标识xls，None表示未取得excel文件
        self.FONT = Font(underline='single', color='FF0000FF')
        self.MainWindow = MainWindow
        self.ui = ui

    # 判断excel是否为xlsx文件
    def isXlsx(self, excelFilePath: str) -> bool:
        if Path(excelFilePath).suffix == '.xlsx':
            self._EXCEL_FLAG = 1
            return True
        elif Path(excelFilePath).suffix == '.xls':
            self._EXCEL_FLAG = 0
            return False
        else:
            self._EXCEL_FLAG = None
            return None

    # 获取Excel路径,并赋给变量工作目录
    def getExcelFilePath(self) -> str:
        excelFile = QFileDialog.getOpenFileName(
            self.MainWindow, "选择excel文件", self._EXCEL_FILE_DIR, 'Excel Files (*.xlsx *.xls)')  # 返回选中的路径形成的列表
        self.isXlsx(excelFile[0])
        self._EXCEL_FILE_DIR = str(Path(excelFile[0]).parent)
        return excelFile[0]

    # 获取Excel的sheet清单
    def getExcelSheetName(self, path: str) -> list:
        if self._EXCEL_FLAG == None:
            print('获取excel的sheet清单无效,未选择excel对象')
            return []
        with open(path, 'rb') as f:
            if self._EXCEL_FLAG:
                wb = openpyxl.load_workbook(f)
                return [i.title for i in wb]
            else:
                return xlrd.open_workbook(file_contents=f.read()).sheet_names()

    # 定义tableWidget控件的排版方法，此控件9*6
    def gridTableWidget(self, list: list) -> None:

        def _ListIter(list):
            for i in list:
                yield i
        listIter = _ListIter(list)
        self.ui.tableWidget.clearContents()
        itemColumn = 0

        while itemColumn < self.ui.tableWidget.columnCount():
            itemRow = 0
            while itemRow < self.ui.tableWidget.rowCount():
                try:
                    self.ui.tableWidget.setItem(
                        itemRow, itemColumn, QTableWidgetItem(next(listIter)))
                    itemRow += 1
                except StopIteration:
                    return None
            itemColumn += 1

    # 定义获取tableWidget控件选定内容的方法
    def getTabArray(self) -> list:
        return [i.text() for i in self.ui.tableWidget.selectedItems()]

    # 定义刷新表格区域的方法
    def refTabArray(self) -> None:
        try:
            excelFilePath = self.ui.lineEdit.text()  # 读取lineEdit存储的excel路径
            wbSheetsList = self.getExcelSheetName(excelFilePath)
            self.ui.tableWidget.setColumnCount(len(wbSheetsList)//9+1)
            # 将sheet清单写入tableWidget
            self.gridTableWidget(wbSheetsList)
        except:
            print(wbSheetsList)
            print('刷新表格出现问题')
        finally:
            pass

    # 定义openFile按钮的动作
    def cmdOpenExcelFile(self) -> None:
        excelFilePath = self.getExcelFilePath()  # 获取excel路径
        if excelFilePath != '':
            self.ui.lineEdit.setText(excelFilePath)  # 将excel路径写入lineEdit
        else:
            return None
        self.refTabArray()

    # 定义commitFileCMD按钮的动作
    def cmdCommitFile(self) -> None:
        if self._EXCEL_FLAG == None:
            print("未选择任何excel对象")
            return None
        excelFilePath = self.ui.lineEdit.text()  # 读取lineEdit存储的excel路径
        with open(excelFilePath, 'rb') as f:
            if self._EXCEL_FLAG:
                wb = openpyxl.load_workbook(f)
                if self.getTabArray() != []:
                    wsWorkSheetList = self.getTabArray()
                else:
                    wsWorkSheetList = [i.title for i in wb if i.title != '目录']

                if '目录' not in [i.title for i in wb]:  # 建立空的目录sheet
                    wb.create_sheet('目录', 0)
                else:
                    wb['目录'].delete_cols(1, 2)
                wsList = wb['目录']

                rownum, colnum = 1, 1
                wsList.cell(row=rownum, column=colnum, value='目 录')
                for i in wsWorkSheetList:
                    rownum += 1
                    wsList.cell(row=rownum, column=colnum,
                                value=i).font = self.FONT
                    wsList.cell(row=rownum, column=colnum).hyperlink = Hyperlink(
                        ref='', location='\'%s\'!A1' % i, tooltip=None, display='%s' % i, id=None)

                    wb[i]['A3'].font = self.FONT
                    wb[i]['A3'].hyperlink = Hyperlink(
                        ref='', location='\'目录\'!A1', tooltip=None, display='目录', id=None)
                wb.save(excelFilePath)
                wb.close()
                self.getTabArray()
                self.refTabArray()
                QMessageBox.information(self.MainWindow, '信息', '建立成功！')
            else:
                pass
